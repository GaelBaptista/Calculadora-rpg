#!/usr/bin/env python3
"""Parse T20 treasure spreadsheet into treasure-data.json"""
import json
import re
from datetime import datetime, date
import openpyxl
from pathlib import Path

XLSX = Path(r'c:\Users\User\Downloads\T20 - Tabela de geração de tesouros.xlsx')
OUT = Path(__file__).resolve().parent.parent / 'treasure-data.json'


def parse_d_range(s):
    """Parse '01-30', '96-100', '100', '11-14' into (min, max)"""
    if s is None:
        return None
    # Excel converte faixas DD-MM em datas (ex: 01-10 → 2025-10-01)
    if isinstance(s, (datetime, date)):
        lo, hi = s.day, s.month
        if lo > hi:
            lo, hi = hi, lo
        return (lo, hi)
    s = str(s).strip()
    if not s or s in ('—', '-', '–'):
        return None
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return None
    s = s.replace(',', '.')
    if '-' in s:
        parts = s.split('-', 1)
        try:
            return (int(float(parts[0])), int(float(parts[1])))
        except ValueError:
            return None
    try:
        v = int(float(s))
        return (v, v)
    except ValueError:
        return None


def parse_nd_key(cell):
    if cell is None:
        return None
    s = str(cell).strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return None
    try:
        v = float(s)
        if v == int(v) and 0 <= v <= 20:
            return str(int(v)) if v >= 1 else '0'
        if v < 1:
            # fractional ND like 0.25 - sheet uses dates for some rows
            return None
    except (ValueError, TypeError):
        pass
    return None


def rows_in_range(entries, roll):
    for e in entries:
        lo, hi = e['range']
        if lo <= roll <= hi:
            return e
    return None


def parse_main_table(ws):
    """Parse 'Tesouro por ND' sheet"""
    nd_tables = {}
    current_nd = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        nd_cell = row[0]
        nd_key = parse_nd_key(nd_cell)
        if nd_key is not None:
            current_nd = nd_key
            if current_nd not in nd_tables:
                nd_tables[current_nd] = {'money': [], 'items': []}

        if current_nd is None:
            continue

        # Na linha do cabeçalho ND, colunas B e E também podem ter entradas
        for col_range, col_result, bucket in [(1, 2, 'money'), (4, 5, 'items')]:
            if len(row) <= col_result:
                continue
            r = parse_d_range(row[col_range])
            res = row[col_result]
            if r and res is not None and str(res).strip():
                entry = {'range': list(r), 'result': str(res).strip()}
                tables = nd_tables[current_nd][bucket]
                if not tables or tables[-1]['range'] != entry['range']:
                    tables.append(entry)

    return nd_tables


def parse_simple_d_table(ws, start_row, col_item=1, col_book=2, col_page=3, skip_headers=True):
    entries = []
    for i, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True)):
        d_range = parse_d_range(row[0])
        if not d_range:
            continue
        item = row[col_item]
        if not item or str(item).strip() in ('d%', 'Item', 'Poção', 'Melhoria', 'Encanto', 'Acessório'):
            continue
        entry = {
            'range': d_range,
            'name': str(item).strip(),
        }
        if col_book < len(row) and row[col_book]:
            entry['book'] = str(row[col_book]).strip()
        if col_page < len(row) and row[col_page]:
            try:
                entry['page'] = int(float(row[col_page]))
            except (ValueError, TypeError):
                pass
        entries.append(entry)
    return entries


def parse_riquezas(ws):
    """Parse wealth tables - menor/media/maior columns"""
    wealth = {
        'menor': [],
        'media': [],
        'maior': [],
        'space': []
    }
    section = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        c0 = str(row[0] or '').strip()
        if c0 == 'VALOR DAS RIQUEZAS':
            section = 'value'
            continue
        if c0 == 'ESPAÇOS DAS RIQUEZAS' or 'ESPA' in c0 and 'RIQUEZAS' in c0:
            section = 'space'
            continue

        if section == 'value':
            r_menor = parse_d_range(row[0])
            r_media = parse_d_range(row[1]) if len(row) > 1 else None
            r_maior = parse_d_range(row[2]) if len(row) > 2 else None
            valor = row[3] if len(row) > 3 else None
            exemplos = row[4] if len(row) > 4 else None

            if r_menor and valor and str(valor) not in ('Menor', 'Valor (T$)'):
                wealth['menor'].append({
                    'range': r_menor,
                    'formula': str(valor).strip(),
                    'examples': str(exemplos).strip() if exemplos else ''
                })
            if r_media and valor and str(row[1] or '') not in ('Média', ''):
                if parse_d_range(row[1]):
                    wealth['media'].append({
                        'range': r_media,
                        'formula': str(valor).strip(),
                        'examples': str(exemplos).strip() if exemplos else ''
                    })
            if r_maior and valor:
                if row[2] and parse_d_range(row[2]) and str(row[2]) not in ('Maior',):
                    wealth['maior'].append({
                        'range': r_maior,
                        'formula': str(valor).strip(),
                        'examples': str(exemplos).strip() if exemplos else ''
                    })

        if section == 'space':
            r = parse_d_range(row[0])
            if r and row[1]:
                try:
                    spaces = float(str(row[1]).replace(',', '.'))
                    wealth['space'].append({
                        'range': r,
                        'spaces': spaces,
                        'desc': str(row[2]).strip() if len(row) > 2 and row[2] else ''
                    })
                except ValueError:
                    pass

    return wealth


def parse_equipment(ws):
    """Three columns: weapons (0-3), armor (5-8), esoteric (10-13)"""
    equip = {'arma': [], 'armadura': [], 'esoterico': []}
    for row in ws.iter_rows(min_row=3, values_only=True):
        # weapons
        r = parse_d_range(row[0])
        if r and row[1]:
            name = str(row[1]).strip()
            if name not in ('Arma', 'd%'):
                equip['arma'].append({'range': r, 'name': name, 'book': str(row[2] or ''), 'page': row[3]})
        # armor
        if len(row) > 5:
            r = parse_d_range(row[5])
            if r and row[6]:
                name = str(row[6]).strip()
                if name not in ('Armadura ou Escudo', 'd%'):
                    equip['armadura'].append({'range': r, 'name': name, 'book': str(row[7] or ''), 'page': row[8]})
        # esoteric
        if len(row) > 10:
            r = parse_d_range(row[10])
            if r and row[11]:
                name = str(row[11]).strip()
                if name not in ('Esotérico', 'd%'):
                    equip['esoterico'].append({'range': r, 'name': name, 'book': str(row[12] or ''), 'page': row[13]})
    return equip


def parse_superiores(ws):
    sup = {'arma': [], 'armadura': [], 'esoterico': []}
    for row in ws.iter_rows(min_row=3, values_only=True):
        r = parse_d_range(row[0])
        if r and row[1]:
            name = str(row[1]).strip()
            if name not in ('Melhoria', 'd%'):
                sup['arma'].append({'range': r, 'name': name})
        if len(row) > 5:
            r = parse_d_range(row[5])
            if r and row[6]:
                name = str(row[6]).strip()
                if name not in ('Melhoria', 'd%'):
                    sup['armadura'].append({'range': r, 'name': name})
        if len(row) > 10:
            r = parse_d_range(row[10])
            if r and row[11]:
                name = str(row[11]).strip()
                if name not in ('Melhoria', 'd%'):
                    sup['esoterico'].append({'range': r, 'name': name})
    return sup


def parse_magicos(ws):
    magic = {
        'arma': [], 'armadura': [], 'esoterico': [],
        'arma_especifica': [], 'armadura_especifica': [], 'esoterico_especifico': []
    }
    mode = 'normal'
    for row in ws.iter_rows(min_row=3, values_only=True):
        c0 = str(row[0] or '').strip()
        if 'ESPECÍFICAS' in c0 or 'ESPEC' in c0:
            if 'ARMAS' in c0:
                mode = 'arma_especifica'
            elif 'ARMADURAS' in c0:
                mode = 'armadura_especifica'
            continue
        if 'ESOTÉRICOS ESPECÍFICOS' in c0 or (c0.startswith('ESOT') and 'ESPEC' in c0):
            mode = 'esoterico_especifico'
            continue

        if mode == 'arma_especifica':
            r = parse_d_range(row[0])
            if r and row[1]:
                magic['arma_especifica'].append({'range': r, 'name': str(row[1]).strip()})
            continue
        if mode == 'armadura_especifica':
            if len(row) > 5:
                r = parse_d_range(row[5])
                if r and row[6]:
                    magic['armadura_especifica'].append({'range': r, 'name': str(row[6]).strip()})
            continue
        if mode == 'esoterico_especifico':
            if len(row) > 10:
                r = parse_d_range(row[10])
                if r and row[11]:
                    magic['esoterico_especifico'].append({'range': r, 'name': str(row[11]).strip()})
            continue

        r = parse_d_range(row[0])
        if r and row[1]:
            name = str(row[1]).strip()
            if name not in ('Encanto', 'd%', 'Arma específica', 'Item específico'):
                magic['arma'].append({'range': r, 'name': name})
        if len(row) > 5:
            r = parse_d_range(row[5])
            if r and row[6]:
                name = str(row[6]).strip()
                if name not in ('Encanto', 'd%'):
                    magic['armadura'].append({'range': r, 'name': name})
        if len(row) > 10:
            r = parse_d_range(row[10])
            if r and row[11]:
                name = str(row[11]).strip()
                if name not in ('Encanto', 'd%', 'Esotérico específico'):
                    magic['esoterico'].append({'range': r, 'name': name})
    return magic


def parse_acessorios(ws):
    acc = {'menor': [], 'medio': [], 'maior': []}
    for row in ws.iter_rows(min_row=3, values_only=True):
        r = parse_d_range(row[0])
        if r and row[1]:
            name = str(row[1]).strip()
            if name not in ('Acessório', 'd%'):
                acc['menor'].append({'range': r, 'name': name, 'price': row[2]})
        if len(row) > 6:
            r = parse_d_range(row[6])
            if r and row[7]:
                name = str(row[7]).strip()
                if name not in ('Acessório', 'd%'):
                    acc['medio'].append({'range': r, 'name': name, 'price': row[8]})
        if len(row) > 12:
            r = parse_d_range(row[12])
            if r and row[13]:
                name = str(row[13]).strip()
                if name not in ('Acessório', 'd%'):
                    acc['maior'].append({'range': r, 'name': name, 'price': row[14]})
    return acc


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # Fix ND 0 from intro - sheet uses date for ND 1/4 etc - read manually from known structure
    nd_table = parse_main_table(wb['Tesouro por ND'])

    # ND fractional (ND < 1) — primeira faixa da planilha
    nd_fractional = {
        '0': {  # ND < 1 (1/4, 1/2 etc use same row as "0" in book - actually first block is ND 1/4)
            'money': [
                {'range': [1, 30], 'result': '—'},
                {'range': [31, 70], 'result': '1d6×10 TC'},
                {'range': [71, 95], 'result': '1d4×100 TC'},
                {'range': [96, 100], 'result': '1d6×10 T$'},
            ],
            'items': [
                {'range': [1, 50], 'result': '—'},
                {'range': [51, 75], 'result': 'Item diverso'},
                {'range': [76, 100], 'result': 'Equipamento'},
            ]
        }
    }

    # Merge - row after dates is ND 1
    if '1' not in nd_table and '1.0' in nd_table:
        nd_table['1'] = nd_table.pop('1.0')
    for k in list(nd_table.keys()):
        nk = str(int(float(k))) if k.replace('.', '').isdigit() else k
        if nk != k:
            nd_table[nk] = nd_table.pop(k)

    nd_table['frac'] = nd_fractional['0']

    data = {
        'version': 'T20-planilha-Guilherme-Dei-Svaldi',
        'ndTable': nd_table,
        'riquezas': parse_riquezas(wb['Riquezas']),
        'itensDiversos': parse_simple_d_table(wb['Itens Diversos'], 3),
        'equipamentos': parse_equipment(wb['Equipamentos']),
        'pocoes': parse_simple_d_table(wb['Poções'], 3, col_item=1, col_book=3, col_page=4),
        'superiores': parse_superiores(wb['Superiores']),
        'magicos': parse_magicos(wb['Mágicos']),
        'acessorios': parse_acessorios(wb['Mágicos (Acessórios)']),
        'rules': {
            'perThreat': True,
            'halfMoney': 'Metade do tesouro: dividir resultados de Dinheiro por 2',
            'double': 'Dobro: role duas vezes em cada coluna',
            'multiCreature': 'Várias criaturas: uma rolagem na linha do ND do encontro'
        }
    }

    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    js_out = OUT.parent / 'treasure-data.js'
    js_out.write_text(
        'window.T20_TREASURE_DATA = ' + json.dumps(data, ensure_ascii=False) + ';\n',
        encoding='utf-8'
    )
    print(f'Wrote {OUT} ({OUT.stat().st_size // 1024} KB)')
    print(f'Wrote {js_out}')
    print('ND keys:', sorted(nd_table.keys(), key=lambda x: (x != 'frac', float(x) if x.replace('.','').isdigit() else 99)))


if __name__ == '__main__':
    main()
